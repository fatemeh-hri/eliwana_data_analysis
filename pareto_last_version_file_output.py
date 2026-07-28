import pandas as pd
import geopandas as gpd
from shapely import wkt
from shapely.geometry import LineString
from sqlalchemy import create_engine
import ezdxf
from functools import lru_cache
import traceback
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dash import Dash, dcc, html, dash_table
from dash.dependencies import Input, Output
import plotly.graph_objects as go
from plotly.subplots import make_subplots


# -----------------------------
# CONFIG
# -----------------------------
DXF_PATH = r"Map_Lane/FMS_Map_20260520_135053.dxf"
INTERSECTION_CSV_PATH = r"final_intersections_after_manual_edits.csv"
EXPORT_XLSX_PATH = r"eliwana_pareto_output.xlsx"

SOURCE_CRS = "EPSG:28350"
TARGET_CRS = "EPSG:4326"
# export_intersections_csv() writes the final GeoDataFrame in TARGET_CRS.
INTERSECTION_CSV_CRS = TARGET_CRS

DEFAULT_START_DATETIME = "2026-07-01 00:00:00"
DEFAULT_END_DATETIME = "2026-07-15 00:00:00"

# DEFAULT_START_DATETIME = "2026-06-01 00:00:00"
# DEFAULT_END_DATETIME = "2026-06-02 00:00:00"

# DEFAULT_START_DATETIME = "2026-06-01 11:00:00"
# DEFAULT_END_DATETIME = "2026-06-01 11:30:00"

LANE_BUFFER_METRES = 10
INTERSECTION_POINT_TOLERANCE_METRES = 10
MAX_INTERACTION_GAP_SECONDS = 5
NEAR_MISS_DISTANCE_METRES = 15
PROCEDURE_BREACH_DISTANCE_METRES = 50

# Configure these patterns to match the actual DXF layer naming convention.
# A matched layer is treated as a narrower LV road; other lane layers are
# treated as haul roads. The road type is contextual only and does not change
# the 15 m / 50 m interaction thresholds.
LV_ROAD_LAYER_PATTERNS = ("LV", "LIGHT_VEHICLE", "LIGHT VEHICLE")

# Historical rule-based risk weights. These are transparent prioritisation
# weights, not a predictive model.
RISK_WEIGHTS = {
    "NEAR_MISS": 5.0,
    "PROCEDURE_BREACH": 2.0,
    "INTERSECTION_MULTIPLIER": 1.5,
    "POTENTIAL_HIERARCHY_BREACH": 2.0,
}

# Uses LEFT(MACHINE, 2) style prefixes
HOV_ASSET_TYPES = ["DZ", "WD", "WC", "GR", "LV", "EL", "WL", "EX"]
# HOV_ASSET_TYPES = ["EX"]
DEBUG = True


# -----------------------------
# DEBUGGING
# -----------------------------
def debug_log(label, value=None):
    if not DEBUG:
        return

    now = datetime.now().strftime("%H:%M:%S")
    print(f"\n[{now}] {label}", flush=True)

    if value is not None:
        if isinstance(value, pd.DataFrame):
            print(f"shape={value.shape}", flush=True)
            print(value.head(), flush=True)
        else:
            print(value, flush=True)


# -----------------------------
# HELPERS
# -----------------------------
def normalise_datetime_for_cache(value):
    return pd.to_datetime(value).strftime("%Y-%m-%d %H:%M:%S")


def truthy_series(series):
    if series is None:
        return pd.Series(dtype=bool)
    return series.astype(str).str.upper().isin(["TRUE", "1", "T", "YES"])


def classify_road_type_from_layer(layer_series):
    """Classify nearest DXF lane as Haul Road or LV Road from its layer name."""
    layer_text = layer_series.fillna("").astype(str).str.upper()
    is_lv = pd.Series(False, index=layer_text.index)
    for pattern in LV_ROAD_LAYER_PATTERNS:
        is_lv |= layer_text.str.contains(pattern, regex=False)
    return pd.Series("Haul Road", index=layer_text.index).where(~is_lv, "LV Road")


def add_historical_risk_analysis(df):
    """Add distance, hierarchy and transparent historical risk classifications."""
    out = df.copy()
    distance = pd.to_numeric(out["DISTANCE_METRES"], errors="coerce")

    out["RISK_EVENT_TYPE"] = "Procedure Breach"
    out.loc[distance < NEAR_MISS_DISTANCE_METRES, "RISK_EVENT_TYPE"] = "Near Miss"
    out.loc[distance > PROCEDURE_BREACH_DISTANCE_METRES, "RISK_EVENT_TYPE"] = "Outside CAS Threshold"

    is_intersection = out["AREA_CATEGORY"].eq("Intersection")
    hov_moving = pd.to_numeric(out["PAIR_IS_MOVING"], errors="coerce").fillna(0).eq(1)
    ht_moving = pd.to_numeric(out["IS_MOVING"], errors="coerce").fillna(0).eq(1)

    out["HIERARCHY_ASSESSMENT"] = "Not Applicable - Straight Path"
    out.loc[is_intersection, "HIERARCHY_ASSESSMENT"] = "Intersection Interaction - Review"
    # out.loc[is_intersection & ht_moving & ~hov_moving, "HIERARCHY_ASSESSMENT"] = (
    #     "HT moving and HOV stationary"
    # )
    # out.loc[is_intersection & hov_moving, "HIERARCHY_ASSESSMENT"] = (
    #     "HOV moving in intersection"
    # )
    out.loc[is_intersection & ht_moving, "HIERARCHY_ASSESSMENT"] = (
        "HT moving in intersection"
    )
    out.loc[is_intersection & ~ht_moving & ~hov_moving, "HIERARCHY_ASSESSMENT"] = (
        "Both Stationary / Queue Review"
    )

    out["IS_POTENTIAL_HIERARCHY_BREACH"] = (
        is_intersection & hov_moving & ht_moving
    )

    base_score = pd.Series(RISK_WEIGHTS["PROCEDURE_BREACH"], index=out.index)
    base_score = base_score.where(
        ~out["RISK_EVENT_TYPE"].eq("Near Miss"),
        RISK_WEIGHTS["NEAR_MISS"],
    )
    base_score = base_score.where(
        ~out["RISK_EVENT_TYPE"].eq("Outside CAS Threshold"),
        0.0,
    )

    out["RISK_EXPOSURE_SCORE"] = base_score
    out.loc[is_intersection, "RISK_EXPOSURE_SCORE"] *= RISK_WEIGHTS["INTERSECTION_MULTIPLIER"]
    out.loc[out["IS_POTENTIAL_HIERARCHY_BREACH"], "RISK_EXPOSURE_SCORE"] += (
        RISK_WEIGHTS["POTENTIAL_HIERARCHY_BREACH"]
    )

    out["RISK_CONTEXT"] = (
        out["AREA_CATEGORY"].astype(str)
        + " | " + out["ROAD_TYPE"].astype(str)
        + " | " + out["RISK_EVENT_TYPE"].astype(str)
    )
    return out


# -----------------------------
# SNOWFLAKE CONNECTION
# -----------------------------
@lru_cache(maxsize=1)
def get_snowflake_engine():
    connection_string = (
        "snowflake://{user}:{password}@{account}/{database}/{schema}"
        "?warehouse={warehouse}&role={role}&authenticator=externalbrowser"
    ).format(
        account=os.environ["SNOWFLAKE_ACCOUNT"],
        user=os.environ["SNOWFLAKE_USER"],
        password = "",
        authenticator=os.getenv("SNOWFLAKE_AUTHENTICATOR", "externalbrowser"),
        role=os.environ["SNOWFLAKE_ROLE"],
        warehouse=os.environ["SNOWFLAKE_WAREHOUSE"],
        database=os.environ["SNOWFLAKE_DATABASE"],
        schema=os.environ["SNOWFLAKE_SCHEMA"],
    )
    return create_engine(connection_string)


# -----------------------------
# DXF LANE LOADER
# -----------------------------
@lru_cache(maxsize=1)
def load_dxf_lanes_source():
    doc = ezdxf.readfile(DXF_PATH)
    msp = doc.modelspace()

    records = []

    for entity in msp.query("LWPOLYLINE"):
        layer_name = entity.dxf.layer.upper()

        if "LANES" not in layer_name:
            continue

        points = [(point[0], point[1]) for point in entity.get_points()]

        if len(points) < 2:
            continue

        records.append({
            "polyline_id": entity.dxf.handle,
            "lane_layer": layer_name,
            "lane_vertex_count": len(points),
            "geometry": LineString(points),
        })

    lanes = gpd.GeoDataFrame(records, geometry="geometry", crs=SOURCE_CRS)

    if lanes.empty:
        return lanes

    lanes = lanes[~lanes.geometry.is_empty].copy()
    lanes["geometry"] = lanes.geometry.simplify(0.3, preserve_topology=True)

    _ = lanes.sindex

    return lanes


# -----------------------------
# INTERSECTION POLYGON LOADER
# -----------------------------
@lru_cache(maxsize=1)
def load_intersection_polygons_source():
    """Load exported intersection polygons and project them to SOURCE_CRS."""
    candidate_paths = [
        INTERSECTION_CSV_PATH,
        os.path.join(os.path.dirname(os.path.abspath(__file__)), INTERSECTION_CSV_PATH),
    ]
    csv_path = next((path for path in candidate_paths if os.path.exists(path)), None)

    if csv_path is None:
        raise FileNotFoundError(
            "Intersection polygon CSV was not found. Expected: "
            + " or ".join(candidate_paths)
            + ". Enable export_intersections_csv(...) in the map workflow first."
        )

    intersection_df = pd.read_csv(csv_path)

    if "geometry_wkt" not in intersection_df.columns:
        raise ValueError(f"{csv_path} must contain a geometry_wkt column")

    if "INTERSECTION_NAME" not in intersection_df.columns:
        intersection_df["INTERSECTION_NAME"] = [
            f"Intersection {i + 1}" for i in range(len(intersection_df))
        ]

    intersection_df = intersection_df.dropna(subset=["geometry_wkt"]).copy()
    intersection_df["geometry"] = intersection_df["geometry_wkt"].map(wkt.loads)

    intersections = gpd.GeoDataFrame(
        intersection_df[["INTERSECTION_NAME", "geometry"]],
        geometry="geometry",
        crs=INTERSECTION_CSV_CRS,
    )
    intersections = intersections[
        intersections.geometry.notna() & ~intersections.geometry.is_empty
    ].copy()

    if intersections.empty:
        raise ValueError(f"{csv_path} contains no valid intersection polygons")

    # Exported polygons are longitude/latitude. Spatial distance and buffering
    # must be performed after projection into MGA Zone 50 metres.
    intersections = intersections.to_crs(SOURCE_CRS)
    intersections["geometry"] = intersections.geometry.buffer(
        INTERSECTION_POINT_TOLERANCE_METRES
    )
    intersections = intersections[
        intersections.geometry.notna() & ~intersections.geometry.is_empty
    ].copy()
    _ = intersections.sindex

    debug_log(
        "Intersection polygons loaded",
        {
            "path": os.path.abspath(csv_path),
            "count": len(intersections),
            "source_crs": INTERSECTION_CSV_CRS,
            "classification_crs": SOURCE_CRS,
            "bounds": intersections.total_bounds.tolist(),
        },
    )
    return intersections


# -----------------------------
# SNOWFLAKE DATA LOADER
# -----------------------------
def load_nearby_machine_data(start_date, end_date):
    hov_type_sql = ", ".join(f"'{asset}'" for asset in HOV_ASSET_TYPES)

    query = f"""
WITH base AS (
    SELECT
        MACHINE,
        "TIMESTAMP",
        X,
        Y,
        Z,
        WKT_GEOM,
        THE_GEOM,
        DATE_TRUNC('SECOND', "TIMESTAMP") AS TS_SECOND,

        LAG(X) OVER (
            PARTITION BY MACHINE
            ORDER BY "TIMESTAMP"
        ) AS PREV_X,

        LAG(Y) OVER (
            PARTITION BY MACHINE
            ORDER BY "TIMESTAMP"
        ) AS PREV_Y,

        LAG(Z) OVER (
            PARTITION BY MACHINE
            ORDER BY "TIMESTAMP"
        ) AS PREV_Z
    FROM AA_OPERATIONS_MANAGEMENT.SELFSERVICE.FMS_MACHINE_LOCATION
    WHERE "TIMESTAMP" >= '{start_date}'
      AND "TIMESTAMP" <  '{end_date}'
),

movement AS (
    SELECT
        *,
        CASE
            WHEN MACHINE LIKE 'ELI%' THEN 'EL'
            WHEN LEFT(MACHINE, 2) = 'LV' THEN 'EL'
            ELSE LEFT(MACHINE, 2)
        END AS ASSET_PREFIX,

        FLOOR(X / 50) AS GRID_X,
        FLOOR(Y / 50) AS GRID_Y,

        CASE
            WHEN PREV_X IS NULL THEN 0
            WHEN SQRT(
                POWER(X - PREV_X, 2) +
                POWER(Y - PREV_Y, 2) +
                POWER(Z - PREV_Z, 2)
            ) > 0.5 THEN 1
            ELSE 0
        END AS IS_MOVING
    FROM base
),

fleet_size AS (
    SELECT
        ASSET_PREFIX,
        COUNT(DISTINCT MACHINE) AS FLEET_SIZE
    FROM movement
    GROUP BY ASSET_PREFIX
),

ht AS (
    SELECT *
    FROM movement
    WHERE ASSET_PREFIX = 'DT'
),

hov AS (
    SELECT *
    FROM movement
    WHERE ASSET_PREFIX IN ({hov_type_sql})
),

raw_matches AS (
    SELECT
        h.MACHINE AS HT_MACHINE,
        l.MACHINE AS HOV_MACHINE,

        h.ASSET_PREFIX AS HT_PREFIX,
        l.ASSET_PREFIX AS HOV_PREFIX,

        h.TS_SECOND,

        ST_DISTANCE(h.THE_GEOM, l.THE_GEOM) AS DISTANCE_METRES,

        h.X AS HT_X,
        h.Y AS HT_Y,
        h.Z AS HT_Z,
        h.WKT_GEOM AS HT_WKT_GEOM,

        l.X AS HOV_X,
        l.Y AS HOV_Y,
        l.Z AS HOV_Z,
        l.WKT_GEOM AS HOV_WKT_GEOM,

        h.IS_MOVING AS HT_IS_MOVING,
        l.IS_MOVING AS HOV_IS_MOVING
    FROM ht h
    JOIN hov l
        ON h.TS_SECOND = l.TS_SECOND
       AND l.GRID_X BETWEEN h.GRID_X - 1 AND h.GRID_X + 1
       AND l.GRID_Y BETWEEN h.GRID_Y - 1 AND h.GRID_Y + 1
    WHERE h.IS_MOVING = 1
       --OR l.IS_MOVING = 1
),

matches AS (
    SELECT
        *,
        TRUE AS IS_VEHICLE_INTERACTION,

        CASE
            WHEN DISTANCE_METRES < 15 THEN TRUE
            ELSE FALSE
        END AS IS_NEAR_MISS,

        'VEHICLE_INTERACTION' AS INTERACTION_TYPE,

        CASE
            WHEN HT_IS_MOVING = 1 AND HOV_IS_MOVING = 1 THEN 'BOTH_MOVING'
            WHEN HT_IS_MOVING = 1 THEN 'HT_DT_MOVING'
            WHEN HOV_IS_MOVING = 1 THEN 'HOV_MOVING'
            ELSE 'UNKNOWN'
        END AS MOVING_STATUS
    FROM raw_matches
    WHERE DISTANCE_METRES <= 50
),

sequenced AS (
    SELECT
        *,
        LAG(TS_SECOND) OVER (
            PARTITION BY HT_MACHINE, HOV_MACHINE
            ORDER BY TS_SECOND, DISTANCE_METRES
        ) AS PREV_TS_SECOND
    FROM matches
),

interaction_events AS (
    SELECT
        *,
        DATEDIFF('second', PREV_TS_SECOND, TS_SECOND) AS TIME_GAP_SECONDS,

        SUM(
            CASE
                WHEN PREV_TS_SECOND IS NULL THEN 1
                WHEN DATEDIFF('second', PREV_TS_SECOND, TS_SECOND) > {MAX_INTERACTION_GAP_SECONDS} THEN 1
                ELSE 0
            END
        ) OVER (
            PARTITION BY HT_MACHINE, HOV_MACHINE
            ORDER BY TS_SECOND, DISTANCE_METRES
            ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
        ) AS INTERACTION_ID,

        CASE
            WHEN PREV_TS_SECOND IS NULL THEN TRUE
            WHEN DATEDIFF('second', PREV_TS_SECOND, TS_SECOND) > {MAX_INTERACTION_GAP_SECONDS} THEN TRUE
            ELSE FALSE
        END AS IS_FIRST_INTERACTION_POINT
    FROM sequenced
)

SELECT
    e.HT_MACHINE,
    e.HOV_MACHINE,

    e.HT_PREFIX,
    e.HOV_PREFIX,

    e.HT_PREFIX || '-' || e.HOV_PREFIX AS ASSET_PAIR,

    ht_fs.FLEET_SIZE AS HT_FLEET_SIZE,
    hov_fs.FLEET_SIZE AS HOV_FLEET_SIZE,

    e.TS_SECOND AS "TIMESTAMP",
    e.DISTANCE_METRES,

    e.IS_VEHICLE_INTERACTION,
    e.IS_NEAR_MISS,
    e.INTERACTION_TYPE,

    e.HT_X,
    e.HT_Y,
    e.HT_Z,
    e.HT_WKT_GEOM,

    e.HOV_X,
    e.HOV_Y,
    e.HOV_Z,
    e.HOV_WKT_GEOM,

    e.HT_IS_MOVING,
    e.HOV_IS_MOVING,
    e.MOVING_STATUS,

    e.TIME_GAP_SECONDS,
    e.INTERACTION_ID,
    e.IS_FIRST_INTERACTION_POINT
FROM interaction_events e
LEFT JOIN fleet_size ht_fs
    ON e.HT_PREFIX = ht_fs.ASSET_PREFIX
LEFT JOIN fleet_size hov_fs
    ON e.HOV_PREFIX = hov_fs.ASSET_PREFIX
WHERE e.IS_FIRST_INTERACTION_POINT = TRUE
ORDER BY "TIMESTAMP", e.DISTANCE_METRES
"""

    debug_log("Snowflake query timeframe", f"{start_date} to {end_date}")

    engine = get_snowflake_engine()
    with engine.connect() as connection:
        df = pd.read_sql(query, connection).rename(columns=str.upper)

    debug_log("Snowflake rows returned", df)

    return df


def load_nearby_machine_data_chunked(start_date, end_date, chunk_hours=6):
    start_dt = pd.to_datetime(start_date)
    end_dt = pd.to_datetime(end_date)

    all_chunks = []
    chunk_start = start_dt

    while chunk_start < end_dt:
        chunk_end = min(chunk_start + pd.Timedelta(hours=chunk_hours), end_dt)

        debug_log("Loading Snowflake chunk", f"{chunk_start} to {chunk_end}")

        df_chunk = load_nearby_machine_data(
            chunk_start.strftime("%Y-%m-%d %H:%M:%S"),
            chunk_end.strftime("%Y-%m-%d %H:%M:%S"),
        )

        if not df_chunk.empty:
            all_chunks.append(df_chunk)

        chunk_start = chunk_end

    if not all_chunks:
        return pd.DataFrame()

    df = pd.concat(all_chunks, ignore_index=True)

    return df.drop_duplicates(
        subset=["HT_MACHINE", "HOV_MACHINE", "TIMESTAMP", "DISTANCE_METRES"]
    )


# -----------------------------
# AREA CLASSIFICATION
# -----------------------------
def classify_area_categories(gdf):
    if gdf.empty:
        return gdf

    lanes_source = load_dxf_lanes_source()
    intersections_source = load_intersection_polygons_source()
    classified = gdf.copy().reset_index(drop=True)
    points_source = classified.to_crs(SOURCE_CRS).copy()
    points_source["SOURCE_ROW_ID"] = classified.index

    # Default category when no lane or intersection is matched.
    classified["DISTANCE_TO_LANE_M"] = pd.NA
    classified["polyline_id"] = pd.NA
    classified["lane_layer"] = pd.NA
    classified["lane_vertex_count"] = pd.NA
    classified["ON_LANE"] = False
    classified["AREA_CATEGORY"] = "Dynamic Area"
    classified["ROAD_TYPE"] = "Off Road / Dynamic Area"
    classified["INTERSECTION_NAME"] = pd.NA

    # Vectorised nearest-lane classification.
    if not lanes_source.empty:
        lane_columns = [
            "polyline_id", "lane_layer", "lane_vertex_count", "geometry"
        ]
        nearest = gpd.sjoin_nearest(
            points_source[["SOURCE_ROW_ID", "geometry"]],
            lanes_source[lane_columns],
            how="left",
            distance_col="DISTANCE_TO_LANE_M",
        )
        nearest = (
            nearest
            .sort_values(["SOURCE_ROW_ID", "DISTANCE_TO_LANE_M"])
            .drop_duplicates(subset="SOURCE_ROW_ID", keep="first")
            .set_index("SOURCE_ROW_ID")
        )

        classified["DISTANCE_TO_LANE_M"] = nearest[
            "DISTANCE_TO_LANE_M"
        ].reindex(classified.index).to_numpy()
        classified["polyline_id"] = nearest[
            "polyline_id"
        ].reindex(classified.index).to_numpy()
        classified["lane_layer"] = nearest[
            "lane_layer"
        ].reindex(classified.index).to_numpy()
        classified["lane_vertex_count"] = nearest[
            "lane_vertex_count"
        ].reindex(classified.index).to_numpy()

        classified["ON_LANE"] = (
            pd.to_numeric(classified["DISTANCE_TO_LANE_M"], errors="coerce")
            <= LANE_BUFFER_METRES
        )
        nearest_road_type = classify_road_type_from_layer(classified["lane_layer"])
        classified.loc[classified["ON_LANE"], "ROAD_TYPE"] = nearest_road_type[classified["ON_LANE"]]
        classified.loc[classified["ON_LANE"], "AREA_CATEGORY"] = "Haul Road"

    # Intersection takes precedence over Haul Road and Dynamic Area.
    if not intersections_source.empty:
        intersection_hits = gpd.sjoin(
            points_source[["SOURCE_ROW_ID", "geometry"]],
            intersections_source[["INTERSECTION_NAME", "geometry"]],
            how="inner",
            predicate="within",
        )

        if not intersection_hits.empty:
            intersection_hits = (
                intersection_hits
                .sort_values("SOURCE_ROW_ID")
                .drop_duplicates(subset="SOURCE_ROW_ID", keep="first")
                .set_index("SOURCE_ROW_ID")
            )
            hit_ids = intersection_hits.index.intersection(classified.index)
            classified.loc[hit_ids, "AREA_CATEGORY"] = "Intersection"
            # Keep nearest-lane ROAD_TYPE to describe the intersection approach.
            classified.loc[hit_ids, "INTERSECTION_NAME"] = (
                intersection_hits.loc[hit_ids, "INTERSECTION_NAME"].to_numpy()
            )

    debug_log(
        "Area classification completed",
        classified["AREA_CATEGORY"].value_counts(dropna=False).to_dict(),
    )
    return classified


# -----------------------------
# PROCESS PAIR DATA
# -----------------------------
def process_pair_data(df):
    if df.empty:
        return gpd.GeoDataFrame()

    df = df.copy()
    df["TIMESTAMP"] = pd.to_datetime(df["TIMESTAMP"])

    df["IS_NEAR_MISS"] = truthy_series(df["IS_NEAR_MISS"])
    df["IS_VEHICLE_INTERACTION"] = truthy_series(df["IS_VEHICLE_INTERACTION"])
    df["IS_FIRST_INTERACTION_POINT"] = truthy_series(df["IS_FIRST_INTERACTION_POINT"])

    df["HT_GEOMETRY"] = df["HT_WKT_GEOM"].apply(wkt.loads)
    df["HOV_GEOMETRY"] = df["HOV_WKT_GEOM"].apply(wkt.loads)

    ht_df = df[[
        "HT_MACHINE", "HT_PREFIX", "HOV_MACHINE", "HOV_PREFIX",
        "TIMESTAMP", "DISTANCE_METRES", "INTERACTION_TYPE", "IS_NEAR_MISS",
        "IS_VEHICLE_INTERACTION", "MOVING_STATUS", "HT_IS_MOVING", "HOV_IS_MOVING",
        "HT_X", "HT_Y", "HT_Z", "HT_GEOMETRY", "HT_FLEET_SIZE", "HOV_FLEET_SIZE",
        "TIME_GAP_SECONDS", "IS_FIRST_INTERACTION_POINT", "INTERACTION_ID", "ASSET_PAIR"
    ]].copy()

    ht_df.columns = [
        "MACHINE", "PREFIX", "PAIR_MACHINE", "PAIR_PREFIX",
        "TIMESTAMP", "DISTANCE_METRES", "INTERACTION_TYPE", "IS_NEAR_MISS",
        "IS_VEHICLE_INTERACTION", "MOVING_STATUS", "IS_MOVING", "PAIR_IS_MOVING",
        "X", "Y", "Z", "geometry", "FLEET_SIZE", "PAIR_FLEET_SIZE",
        "TIME_GAP_SECONDS", "IS_FIRST_INTERACTION_POINT", "INTERACTION_ID", "ASSET_PAIR"
    ]
    ht_df["TYPE"] = "HT"

    hov_df = df[[
        "HOV_MACHINE", "HOV_PREFIX", "HT_MACHINE", "HT_PREFIX",
        "TIMESTAMP", "DISTANCE_METRES", "INTERACTION_TYPE", "IS_NEAR_MISS",
        "IS_VEHICLE_INTERACTION", "MOVING_STATUS", "HOV_IS_MOVING", "HT_IS_MOVING",
        "HOV_X", "HOV_Y", "HOV_Z", "HOV_GEOMETRY", "HOV_FLEET_SIZE", "HT_FLEET_SIZE",
        "TIME_GAP_SECONDS", "IS_FIRST_INTERACTION_POINT", "INTERACTION_ID", "ASSET_PAIR"
    ]].copy()

    hov_df.columns = [
        "MACHINE", "PREFIX", "PAIR_MACHINE", "PAIR_PREFIX",
        "TIMESTAMP", "DISTANCE_METRES", "INTERACTION_TYPE", "IS_NEAR_MISS",
        "IS_VEHICLE_INTERACTION", "MOVING_STATUS", "IS_MOVING", "PAIR_IS_MOVING",
        "X", "Y", "Z", "geometry", "FLEET_SIZE", "PAIR_FLEET_SIZE",
        "TIME_GAP_SECONDS", "IS_FIRST_INTERACTION_POINT", "INTERACTION_ID", "ASSET_PAIR"
    ]
    hov_df["TYPE"] = "HOV"

    plot_df = pd.concat([ht_df, hov_df], ignore_index=True)

    plot_df = plot_df.drop_duplicates(
        subset=[
            "TYPE", "PREFIX", "PAIR_PREFIX", "MACHINE",
            "PAIR_MACHINE", "TIMESTAMP", "X", "Y"
        ]
    )

    gdf = gpd.GeoDataFrame(plot_df, geometry="geometry", crs=SOURCE_CRS).to_crs(TARGET_CRS)

    gdf["lon"] = gdf.geometry.x
    gdf["lat"] = gdf.geometry.y
    gdf["TIMESTAMP_STR"] = gdf["TIMESTAMP"].astype(str).str.split("+").str[0]
    gdf["TIME_INT"] = (
        gdf["TIMESTAMP"] - gdf["TIMESTAMP"].min()
    ).dt.total_seconds().astype(int)

    debug_log("process_pair_data output", f"{len(gdf)} rows")

    return classify_area_categories(gdf)


# -----------------------------
# PARETO SUMMARY HELPERS
# -----------------------------
def summarise_for_pareto(pareto_df, group_columns):
    if pareto_df.empty:
        return pd.DataFrame(
            columns=[
                *group_columns,
                "CATEGORY",
                "VEHICLE_INTERACTIONS",
                "NEAR_MISSES",
                "CUMULATIVE_COUNT",
                "CUMULATIVE_PERCENT",
            ]
        )

    summary = (
        pareto_df
        .groupby(group_columns)
        .agg(
            VEHICLE_INTERACTIONS=("IS_VEHICLE_INTERACTION", "sum"),
            NEAR_MISSES=("IS_NEAR_MISS", "sum"),
        )
        .reset_index()
        .sort_values("VEHICLE_INTERACTIONS", ascending=False)
        .reset_index(drop=True)
    )

    summary["VEHICLE_INTERACTIONS"] = summary["VEHICLE_INTERACTIONS"].astype(int)
    summary["NEAR_MISSES"] = summary["NEAR_MISSES"].astype(int)
    summary["CATEGORY"] = summary[group_columns].astype(str).agg(" | ".join, axis=1)
    summary["CUMULATIVE_COUNT"] = summary["VEHICLE_INTERACTIONS"].cumsum()
    summary["CUMULATIVE_PERCENT"] = (
        summary["CUMULATIVE_COUNT"]
        / summary["VEHICLE_INTERACTIONS"].sum()
        * 100
    )

    return summary


def build_rate_summary(pareto_df, hours):
    if pareto_df.empty:
        return pd.DataFrame(columns=[
            "HOV_ASSET", "FLEET_SIZE", "VEHICLE_INTERACTIONS", "NEAR_MISSES",
            "HOURS", "INTERACTIONS_PER_HOUR", "INTERACTIONS_PER_HOUR_PER_EQUIPMENT"
        ])

    rate = (
        pareto_df
        .groupby("HOV_ASSET")
        .agg(
            FLEET_SIZE=("PAIR_FLEET_SIZE", "max"),
            VEHICLE_INTERACTIONS=("IS_VEHICLE_INTERACTION", "sum"),
            NEAR_MISSES=("IS_NEAR_MISS", "sum"),
        )
        .reset_index()
    )

    rate["FLEET_SIZE"] = rate["FLEET_SIZE"].fillna(1).replace(0, 1).astype(int)
    rate["VEHICLE_INTERACTIONS"] = rate["VEHICLE_INTERACTIONS"].astype(int)
    rate["NEAR_MISSES"] = rate["NEAR_MISSES"].astype(int)
    rate["HOURS"] = hours
    rate["INTERACTIONS_PER_HOUR"] = rate["VEHICLE_INTERACTIONS"] / hours
    rate["INTERACTIONS_PER_HOUR_PER_EQUIPMENT"] = (
        rate["VEHICLE_INTERACTIONS"] / hours / rate["FLEET_SIZE"]
    )

    return rate.sort_values("INTERACTIONS_PER_HOUR_PER_EQUIPMENT", ascending=False)


def prepare_pareto_source(detail_records, hide_loading_operations):
    if not detail_records:
        return pd.DataFrame()

    out = pd.DataFrame(detail_records)
    out["IS_NEAR_MISS"] = out["IS_NEAR_MISS"].astype(bool)
    out["IS_VEHICLE_INTERACTION"] = out["IS_VEHICLE_INTERACTION"].astype(bool)

    if hide_loading_operations:
        out = out[~out["IS_LOADING_OPERATION"]].copy()

    return out


def get_pareto_payload(start_date, end_date):
    start_date = normalise_datetime_for_cache(start_date)
    end_date = normalise_datetime_for_cache(end_date)

    debug_log("get_pareto_payload timeframe", f"{start_date} to {end_date}")

    df = load_nearby_machine_data_chunked(
        start_date,
        end_date,
        chunk_hours=6,
    )

    gdf_plot = process_pair_data(df)

    if not gdf_plot.empty:
        debug_log(
            "Pareto area categories before payload",
            gdf_plot["AREA_CATEGORY"].value_counts(dropna=False).to_dict(),
        )

    if gdf_plot.empty:
        return {
            "detail": [],
            "start_date": start_date,
            "end_date": end_date,
        }

    out = pd.DataFrame(gdf_plot.drop(columns="geometry"))

    out["DT_HT_ASSET"] = out["PREFIX"].where(out["TYPE"].eq("HT"), out["PAIR_PREFIX"])
    out["HOV_ASSET"] = out["PAIR_PREFIX"].where(out["TYPE"].eq("HT"), out["PREFIX"])
    out["AREA_CATEGORY"] = out["AREA_CATEGORY"].fillna("Unknown")
    out["ASSET_PAIR"] = out["DT_HT_ASSET"].astype(str) + "-" + out["HOV_ASSET"].astype(str)
    out["ASSET_AREA_CATEGORY"] = out["ASSET_PAIR"] + " | " + out["AREA_CATEGORY"]

    # Loading operations to exclude from Pareto
    # EX and WL interactions occurring in Dynamic Areas or Haul Roads
    out["IS_LOADING_OPERATION"] = (
        out["TYPE"].eq("HT")
        & out["HOV_ASSET"].isin(["EX", "WL"])
        & out["AREA_CATEGORY"].isin([
            "Dynamic Area",
            "Haul Road",
            "Intersection",
        ])
    )

    pareto_detail = out[out["TYPE"].eq("HT")].copy()
    pareto_detail = add_historical_risk_analysis(pareto_detail)

    keep_columns = [
        "TIMESTAMP_STR", "HT_MACHINE" if "HT_MACHINE" in pareto_detail.columns else "MACHINE",
        "MACHINE", "PAIR_MACHINE", "DT_HT_ASSET", "HOV_ASSET", "ASSET_PAIR",
        "AREA_CATEGORY", "ROAD_TYPE", "ASSET_AREA_CATEGORY", "DISTANCE_METRES",
        "RISK_EVENT_TYPE", "HIERARCHY_ASSESSMENT", "IS_POTENTIAL_HIERARCHY_BREACH",
        "RISK_EXPOSURE_SCORE", "RISK_CONTEXT", "IS_NEAR_MISS",
        "IS_VEHICLE_INTERACTION", "MOVING_STATUS", "IS_MOVING", "PAIR_IS_MOVING",
        "FLEET_SIZE", "PAIR_FLEET_SIZE", "IS_LOADING_OPERATION",
        "DISTANCE_TO_LANE_M", "polyline_id", "lane_layer", "INTERSECTION_NAME"
    ]
    keep_columns = [c for c in keep_columns if c in pareto_detail.columns]

    payload = {
        "detail": pareto_detail[keep_columns].to_dict("records"),
        "start_date": start_date,
        "end_date": end_date,
    }

    debug_log("get_pareto_payload completed", {"detail_rows": len(payload["detail"])})
    return payload


def summarise_risk_exposure(pareto_df, group_columns):
    if pareto_df.empty:
        return pd.DataFrame()

    summary = (
        pareto_df.groupby(group_columns, dropna=False)
        .agg(
            VEHICLE_INTERACTIONS=("IS_VEHICLE_INTERACTION", "sum"),
            NEAR_MISSES=("IS_NEAR_MISS", "sum"),
            POTENTIAL_HIERARCHY_BREACHES=("IS_POTENTIAL_HIERARCHY_BREACH", "sum"),
            RISK_EXPOSURE_SCORE=("RISK_EXPOSURE_SCORE", "sum"),
            MIN_DISTANCE_METRES=("DISTANCE_METRES", "min"),
        )
        .reset_index()
        .sort_values(["RISK_EXPOSURE_SCORE", "NEAR_MISSES"], ascending=False)
        .reset_index(drop=True)
    )
    summary["CATEGORY"] = summary[group_columns].astype(str).agg(" | ".join, axis=1)
    return summary


def build_risk_exposure_chart(summary_records, title):
    summary = pd.DataFrame(summary_records)
    if summary.empty:
        return build_pareto_chart_from_summary([], title)

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=summary["CATEGORY"],
        y=summary["RISK_EXPOSURE_SCORE"],
        name="Historical Risk Exposure Score",
        text=summary["RISK_EXPOSURE_SCORE"].round(1),
        textposition="outside",
        customdata=summary[[
            "VEHICLE_INTERACTIONS", "NEAR_MISSES",
            "POTENTIAL_HIERARCHY_BREACHES", "MIN_DISTANCE_METRES"
        ]],
        hovertemplate=(
            "%{x}<br>Risk score: %{y:.1f}"
            "<br>Interactions: %{customdata[0]}"
            "<br>Near misses: %{customdata[1]}"
            "<br>Potential hierarchy breaches: %{customdata[2]}"
            "<br>Minimum distance: %{customdata[3]:.2f} m<extra></extra>"
        ),
    ))
    fig.update_layout(
        title=title, height=620, margin=dict(l=50, r=50, t=80, b=170),
        xaxis_title="Intersection / Asset Pair", yaxis_title="Historical Risk Exposure Score",
    )
    fig.update_xaxes(tickangle=45)
    return fig


# -----------------------------
# FIGURE HELPERS
# -----------------------------
def build_pareto_chart_from_summary(summary_records, title):
    summary = pd.DataFrame(summary_records)

    if summary.empty:
        fig = go.Figure()
        fig.update_layout(
            title=title,
            height=520,
            annotations=[{
                "text": "No data available for the selected filters",
                "xref": "paper",
                "yref": "paper",
                "x": 0.5,
                "y": 0.5,
                "showarrow": False,
                "font": {"size": 16},
            }],
        )
        return fig

    fig = make_subplots(specs=[[{"secondary_y": True}]])

    fig.add_trace(
        go.Bar(
            x=summary["CATEGORY"],
            y=summary["VEHICLE_INTERACTIONS"],
            name="Vehicle Interactions",
            text=summary["VEHICLE_INTERACTIONS"],
            textposition="outside",
        ),
        secondary_y=False,
    )

    fig.add_trace(
        go.Scatter(
            x=summary["CATEGORY"],
            y=summary["CUMULATIVE_PERCENT"],
            name="Cumulative %",
            mode="lines+markers+text",
            text=summary["CUMULATIVE_PERCENT"].round(1).astype(str) + "%",
            textposition="top center",
        ),
        secondary_y=True,
    )

    fig.update_layout(
        title=title,
        height=620,
        hovermode="x unified",
        margin=dict(l=50, r=50, t=80, b=150),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1,
        ),
    )

    fig.update_xaxes(title_text="Category", tickangle=45)
    fig.update_yaxes(title_text="Vehicle Interactions", secondary_y=False)
    fig.update_yaxes(title_text="Cumulative %", range=[0, 105], secondary_y=True)

    return fig


def write_dataframe_to_worksheet(workbook, sheet_name, dataframe):
    """Write a DataFrame to a formatted Excel worksheet."""
    safe_name = str(sheet_name)[:31]
    worksheet = workbook.create_sheet(title=safe_name)

    if dataframe is None or dataframe.empty:
        worksheet["A1"] = "No data available for the selected filters"
        worksheet["A1"].font = Font(bold=True)
        return worksheet

    export_df = dataframe.copy()

    # Excel cannot store timezone-aware datetimes.
    for column in export_df.columns:
        if pd.api.types.is_datetime64_any_dtype(export_df[column]):
            try:
                export_df[column] = export_df[column].dt.tz_localize(None)
            except TypeError:
                pass

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for col_index, column_name in enumerate(export_df.columns, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=str(column_name))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(export_df.itertuples(index=False, name=None), start=2):
        for col_index, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            elif hasattr(value, "item"):
                try:
                    value = value.item()
                except (ValueError, TypeError):
                    pass
            worksheet.cell(row=row_index, column=col_index, value=value)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for col_index, column_name in enumerate(export_df.columns, start=1):
        values = [str(column_name)]
        values.extend(
            "" if value is None else str(value)
            for value in export_df.iloc[:, col_index - 1].head(1000)
        )
        width = min(max(len(value) for value in values) + 2, 45)
        worksheet.column_dimensions[get_column_letter(col_index)].width = max(width, 12)

    return worksheet


def export_pareto_excel_file(payload, hide_loading_operations, output_path=EXPORT_XLSX_PATH):
    """Export graph data to one Excel workbook with a worksheet per graph."""
    pareto_df = prepare_pareto_source(
        (payload or {}).get("detail", []),
        hide_loading_operations=hide_loading_operations,
    )

    start_dt = pd.to_datetime((payload or {}).get("start_date", DEFAULT_START_DATETIME))
    end_dt = pd.to_datetime((payload or {}).get("end_date", DEFAULT_END_DATETIME))
    hours = max((end_dt - start_dt).total_seconds() / 3600, 1)

    combined = summarise_for_pareto(pareto_df, ["ASSET_PAIR", "AREA_CATEGORY"])
    near_miss = summarise_for_pareto(
        pareto_df[pareto_df["IS_NEAR_MISS"]] if not pareto_df.empty else pareto_df,
        ["ASSET_PAIR", "AREA_CATEGORY"],
    )
    asset = summarise_for_pareto(pareto_df, ["ASSET_PAIR"])
    area = summarise_for_pareto(pareto_df, ["AREA_CATEGORY"])

    intersection_df = (
        pareto_df[pareto_df["AREA_CATEGORY"].eq("Intersection")].copy()
        if not pareto_df.empty
        else pareto_df
    )
    intersection_risk = summarise_risk_exposure(
        intersection_df, ["INTERSECTION_NAME", "ASSET_PAIR"]
    )
    rate_df = build_rate_summary(pareto_df, hours).round(2)

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_dataframe_to_worksheet(workbook, "Combined Pareto", combined)
    write_dataframe_to_worksheet(workbook, "Near Miss Pareto", near_miss)
    write_dataframe_to_worksheet(workbook, "Asset Pareto", asset)
    write_dataframe_to_worksheet(workbook, "Area Pareto", area)
    write_dataframe_to_worksheet(workbook, "Intersection Risk", intersection_risk)
    write_dataframe_to_worksheet(workbook, "Rate Metrics", rate_df)
    write_dataframe_to_worksheet(workbook, "Interaction Detail", pareto_df)

    output_path = os.path.abspath(output_path)
    workbook.save(output_path)
    debug_log("Pareto Excel export completed", output_path)
    return output_path


def build_summary_table_from_records(summary_records):
    if not summary_records:
        return []

    summary = pd.DataFrame(summary_records)

    header_style = {
        "border": "1px solid #ddd",
        "padding": "6px",
        "textAlign": "left",
        "backgroundColor": "#f3f3f3",
    }
    cell_style = {
        "border": "1px solid #ddd",
        "padding": "6px",
    }

    return [
        html.Table(
            children=[
                html.Thead(
                    html.Tr([
                        html.Th("Rank", style=header_style),
                        html.Th("Asset Pair / Area", style=header_style),
                        html.Th("Vehicle Interactions", style=header_style),
                        html.Th("Near Misses", style=header_style),
                        html.Th("Cumulative %", style=header_style),
                    ])
                ),
                html.Tbody([
                    html.Tr([
                        html.Td(i + 1, style=cell_style),
                        html.Td(row["CATEGORY"], style=cell_style),
                        html.Td(int(row["VEHICLE_INTERACTIONS"]), style=cell_style),
                        html.Td(int(row["NEAR_MISSES"]), style=cell_style),
                        html.Td(f"{row['CUMULATIVE_PERCENT']:.1f}%", style=cell_style),
                    ])
                    for i, row in summary.iterrows()
                ]),
            ],
            style={
                "width": "100%",
                "borderCollapse": "collapse",
                "fontSize": "13px",
            },
        )
    ]


def make_card(title, value):
    return html.Div(
        children=[
            html.Div(title, style={"fontSize": "13px", "color": "#555"}),
            html.Div(str(value), style={"fontSize": "22px", "fontWeight": "bold"}),
        ],
        style={
            "border": "1px solid #ddd",
            "borderRadius": "8px",
            "padding": "12px",
            "backgroundColor": "#fafafa",
        },
    )


# -----------------------------
# DASH APP
# -----------------------------
app = Dash(__name__)

app.layout = html.Div(
    style={"padding": "15px", "fontFamily": "Arial, sans-serif"},
    children=[
        dcc.Store(id="pareto-summary-store", storage_type="memory"),

        html.Div(
            "Eliwana HT/DT/HOV Interaction Pareto",
            style={
                "fontSize": "24px",
                "fontWeight": "bold",
                "marginBottom": "15px",
            },
        ),

        dcc.Interval(
            id="initial-load",
            interval=100,
            n_intervals=0,
            max_intervals=1,
        ),

        html.Div(
            id="load-status",
            style={"marginBottom": "15px", "color": "#555"},
        ),

        dcc.Checklist(
            id="hide-loading-toggle",
            options=[{
                "label": "Hide loading operation interactions: EX and WL",
                "value": "hide_loading",
            }],
            value=[],
            style={"marginBottom": "15px"},
        ),

        html.Div(
            id="summary-cards",
            style={
                "display": "grid",
                "gridTemplateColumns": "repeat(auto-fit, minmax(180px, 1fr))",
                "gap": "12px",
                "marginBottom": "20px",
            },
        ),

        dcc.Graph(id="combined-pareto"),
        dcc.Graph(id="near-miss-pareto"),
        dcc.Graph(id="asset-pareto"),
        dcc.Graph(id="area-pareto"),
        dcc.Graph(id="intersection-risk-pareto"),
        # dcc.Graph(id="hierarchy-pareto"),

        html.Div(
            "Interaction Rate Metrics",
            style={
                "fontSize": "18px",
                "fontWeight": "bold",
                "marginTop": "20px",
            },
        ),
        dash_table.DataTable(
            id="rate-table",
            page_size=15,
            style_table={"overflowX": "auto"},
            style_cell={"textAlign": "left", "padding": "6px", "fontSize": 13},
            style_header={"fontWeight": "bold", "backgroundColor": "#f3f3f3"},
        ),

        # html.Div(
        #     "EX Interactions on Haul Road Review",
        #     style={
        #         "fontSize": "18px",
        #         "fontWeight": "bold",
        #         "marginTop": "20px",
        #     },
        # ),
        # dash_table.DataTable(
        #     id="ex-haul-road-table",
        #     page_size=10,
        #     style_table={"overflowX": "auto"},
        #     style_cell={"textAlign": "left", "padding": "6px", "fontSize": 13},
        #     style_header={"fontWeight": "bold", "backgroundColor": "#f3f3f3"},
        # ),

        html.Div(
            "Pareto Detail Table",
            style={
                "fontSize": "18px",
                "fontWeight": "bold",
                "marginTop": "20px",
            },
        ),

        html.Div(id="pareto-table", style={"marginTop": "10px"}),
    ],
)


# -----------------------------
# LOAD DATA ONCE
# -----------------------------
@app.callback(
    Output("pareto-summary-store", "data"),
    Output("load-status", "children"),
    Input("initial-load", "n_intervals"),
)
def load_data_to_store(n_intervals):
    debug_log("load_data_to_store called", n_intervals)

    start_date = normalise_datetime_for_cache(DEFAULT_START_DATETIME)
    end_date = normalise_datetime_for_cache(DEFAULT_END_DATETIME)

    try:
        payload = get_pareto_payload(start_date, end_date)
    except Exception as exc:
        traceback.print_exc()
        return {}, f"Data load failed: {exc}"

    status = f"Loaded Pareto summaries for {start_date} to {end_date}."

    return payload, status


# -----------------------------
# UPDATE PARETO CHARTS
# -----------------------------
@app.callback(
    Output("combined-pareto", "figure"),
    Output("near-miss-pareto", "figure"),
    Output("asset-pareto", "figure"),
    Output("area-pareto", "figure"),
    Output("intersection-risk-pareto", "figure"),
    # Output("hierarchy-pareto", "figure"),
    Output("summary-cards", "children"),
    Output("rate-table", "data"),
    Output("rate-table", "columns"),
    # Output("ex-haul-road-table", "data"),
    # Output("ex-haul-road-table", "columns"),
    Output("pareto-table", "children"),
    Input("pareto-summary-store", "data"),
    Input("hide-loading-toggle", "value"),
)
def update_pareto(payload, hide_loading_value):
    debug_log("update_pareto called")

    hide_loading_operations = "hide_loading" in (hide_loading_value or [])

    if not payload:
        empty_fig = build_pareto_chart_from_summary([], "No data")
        return empty_fig, empty_fig, empty_fig, empty_fig, empty_fig, [], [], [], []

    pareto_df = prepare_pareto_source(
        payload.get("detail", []),
        hide_loading_operations=hide_loading_operations,
    )

    if pareto_df.empty:
        empty_fig = build_pareto_chart_from_summary([], "No data")
        return empty_fig, empty_fig, empty_fig, empty_fig, empty_fig, [], [], [], []

    start_dt = pd.to_datetime(payload.get("start_date", DEFAULT_START_DATETIME))
    end_dt = pd.to_datetime(payload.get("end_date", DEFAULT_END_DATETIME))
    hours = max((end_dt - start_dt).total_seconds() / 3600, 1)

    combined = summarise_for_pareto(pareto_df, ["ASSET_PAIR", "AREA_CATEGORY"])
    asset = summarise_for_pareto(pareto_df, ["ASSET_PAIR"])
    area = summarise_for_pareto(pareto_df, ["AREA_CATEGORY"])
    near_miss = summarise_for_pareto(pareto_df[pareto_df["IS_NEAR_MISS"]], ["ASSET_PAIR", "AREA_CATEGORY"])
    intersection_df = pareto_df[pareto_df["AREA_CATEGORY"].eq("Intersection")].copy()
    intersection_risk = summarise_risk_exposure(
        intersection_df, ["INTERSECTION_NAME", "ASSET_PAIR"]
    )
    hierarchy = summarise_risk_exposure(
        intersection_df, ["HIERARCHY_ASSESSMENT"]
    )

    top_category = combined.iloc[0]["CATEGORY"] if not combined.empty else "-"

    combined_fig = build_pareto_chart_from_summary(
        combined.to_dict("records"),
        "Pareto by DT/HT-HOV Asset Type and Area Category",
    )

    near_miss_fig = build_pareto_chart_from_summary(
        near_miss.to_dict("records"),
        "Near Miss Pareto by DT/HT-HOV Asset Type and Area Category",
    )

    asset_fig = build_pareto_chart_from_summary(
        asset.to_dict("records"),
        "Pareto by DT/HT-HOV Asset Type",
    )

    area_fig = build_pareto_chart_from_summary(
        area.to_dict("records"),
        "Pareto by Dynamic Area / Haul Road / Intersection",
    )

    intersection_risk_fig = build_risk_exposure_chart(
        intersection_risk.to_dict("records"),
        "Historical Intersection Risk Exposure by Intersection and Asset Pair",
    )
    # hierarchy_fig = build_risk_exposure_chart(
    #     hierarchy.to_dict("records"),
    #     "Intersection Hierarchy Review",
    # )

    # ex_haul_road = pareto_df[
    #     pareto_df["HOV_ASSET"].eq("EX")
    #     & pareto_df["AREA_CATEGORY"].eq("Haul Road")
    # ].copy()

    summary_cards = [
        make_card("Vehicle Interactions", f"{int(pareto_df['IS_VEHICLE_INTERACTION'].sum()):,}"),
        make_card("Near Misses", f"{int(pareto_df['IS_NEAR_MISS'].sum()):,}"),
        make_card("Intersection Interactions", f"{len(intersection_df):,}"),
        make_card(
            "Potential Hierarchy Breaches",
            f"{int(intersection_df['IS_POTENTIAL_HIERARCHY_BREACH'].sum()) if not intersection_df.empty else 0:,}",
        ),
        make_card("Top Pareto Category", top_category),
        # make_card("EX on Haul Road", f"{len(ex_haul_road):,}"),
    ]

    rate_df = build_rate_summary(pareto_df, hours).round(2)
    rate_columns = [{"name": c, "id": c} for c in rate_df.columns]

    # ex_columns_selected = [
    #     "TIMESTAMP_STR", "MACHINE", "PAIR_MACHINE", "ASSET_PAIR", "AREA_CATEGORY",
    #     "DISTANCE_METRES", "MOVING_STATUS", "DISTANCE_TO_LANE_M", "polyline_id", "lane_layer"
    # ]
    # ex_columns_selected = [c for c in ex_columns_selected if c in ex_haul_road.columns]
    # ex_table = ex_haul_road[ex_columns_selected].sort_values("TIMESTAMP_STR", ascending=False).round(2)
    # ex_columns = [{"name": c, "id": c} for c in ex_table.columns]

    table_children = build_summary_table_from_records(combined.to_dict("records"))

    # Automatically overwrite the Excel export whenever the dashboard output
    # is recalculated, including changes to the loading-operation filter.
    try:
        export_pareto_excel_file(
            payload,
            hide_loading_operations=hide_loading_operations,
            output_path=EXPORT_XLSX_PATH,
        )
    except Exception as exc:
        traceback.print_exc()
        debug_log("Pareto Excel export failed", str(exc))

    return (
        combined_fig,
        near_miss_fig,
        asset_fig,
        area_fig,
        intersection_risk_fig,
        # hierarchy_fig,
        summary_cards,
        rate_df.to_dict("records"),
        rate_columns,
        # ex_table.to_dict("records"),
        # ex_columns,
        table_children,
    )


# -----------------------------
# RUN
# -----------------------------
if __name__ == "__main__":
    get_snowflake_engine.cache_clear()
    load_dxf_lanes_source.cache_clear()
    load_intersection_polygons_source.cache_clear()

    app.run(debug=True, port=8005, use_reloader=False)
